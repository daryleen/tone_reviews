from superSort import *
import openpyxl
import numpy as np


def create_weights(dataset_word):
    random.seed(2)
    random_weight = []
    for i in range(len(dataset_word)):
        random_weight.append(int(random.random() * 9 + 1))
    return random_weight


def sigmoid(x):
    return 1 / (1 + np.exp(-x))


def sigmoid_derivative(x):
    return x * (1 - x)

pos = neg = false_pos = false_neg = False
net = 0
random_weights = create_weights(dataset_words)
# print(*random_weights)

workbook = openpyxl.load_workbook("результат.xlsx")
sheet = workbook.active

# Загружаем Excel-файл
workbook = openpyxl.load_workbook("результат.xlsx")
sheet = workbook.active

# Считываем данные из ячеек A3 до A502
reviews = []
for row in sheet.iter_rows(min_row=3, max_row=502, min_col=1, max_col=1, values_only=True):
    reviews.append(row[0])

# Считываем значения из ячеек B3 до HH3 и записываем в массив arr
x_values = []
for row in sheet.iter_rows(min_row=3, max_row=3, min_col=2, max_col=249, values_only=True):
    for cell in row:
        if cell is not None:
            x_values.append(cell)

x_values = np.array(x_values)

input_size = len(x_values)
hidden_size = 64
output_size = 1
learning_rate = 0.1

weights_input_hidden = np.random.rand(input_size, hidden_size)
weights_hidden_output = np.random.rand(hidden_size, output_size)
EPOCH = 30
for epoch in range(EPOCH):
    correct_ans = 0
    faults = 0
    shuffled_indices = list(range(len(reviews)))
    for j in shuffled_indices:
        # Установка значения переменной review_tone в зависимости от положения отзыва
        if reviews[j] in reviews[:250]:
            review_tone = 1
        else:
            review_tone = 0
        net = 0

        # Прямое распространение
        hidden_input = np.dot(x_values, weights_input_hidden)
        hidden_output = sigmoid(hidden_input)
        output = np.dot(hidden_output, weights_hidden_output)
        predicted = sigmoid(output)

        # Вычисление ошибки
        error = review_tone - predicted

        # Обратное распространение
        d_predicted = error * sigmoid_derivative(predicted)
        error_hidden = d_predicted.dot(weights_hidden_output.T)
        d_hidden = error_hidden * sigmoid_derivative(hidden_output)

        # Обновление весов
        weights_hidden_output += hidden_output.reshape(-1, 1) * d_predicted * learning_rate
        weights_input_hidden += x_values.reshape(-1, 1) * d_hidden * learning_rate

        # Подсчет правильных и неправильных ответов
        if (review_tone == 1 and predicted > 0.5) or (review_tone == 0 and predicted <= 0.5):
            correct_ans += 1
        else:
            faults += 1

    print(f"Epoch {epoch+1}: Correct Answers: {correct_ans}, Faults: {faults}")