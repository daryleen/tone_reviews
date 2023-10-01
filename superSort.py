import openpyxl
import random


def impoport(name):
    # Открываем файл Excel
    workbook = openpyxl.load_workbook(name)
    # Выбираем активный лист (первый лист в книге)
    sheet = workbook.active

    # Создаем список для хранения текста из каждой ячейки
    cell_text_list = []

    # Проходим по каждой строке и столбцу в таблице
    for row in sheet.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value is not None:
                cell_text_list.append(str(cell_value))

    # Закрываем файл Excel
    workbook.close()
    return cell_text_list


revpos = "positive.xlsx"
revneg = "negative.xlsx"
datawords = "уникальные_слова.xlsx"
# Создаем новую книгу Excel
workbook = openpyxl.Workbook()

# Выбираем активный лист (первый лист в книге)
sheet = workbook.active

# Массивы с отзывами и словами для датасета
positive_reviews = impoport(revpos)
negative_reviews = impoport(revneg)
dataset_words = impoport(datawords)

# Добавляем слова из датасета во вторую строку, начиная со второй ячейки
for col, word in enumerate(dataset_words, start=2):
    sheet.cell(row=2, column=col, value=word)

# Добавляем положительные отзывы в столбец A, начиная со строки 3
for row, review in enumerate(positive_reviews, start=3):
    sheet.cell(row=row, column=1, value=review)

# Добавляем отрицательные отзывы в столбец A, начиная со следующей строки
for row, review in enumerate(negative_reviews, start=len(positive_reviews) + 3):
    sheet.cell(row=row, column=1, value=review)

# Проходим по каждому отзыву и анализируем его на наличие слов из датасета
for row in range(3, len(positive_reviews) + len(negative_reviews) + 3):
    for col in range(2, len(dataset_words) + 2):
        review = sheet.cell(row=row, column=1).value
        word = sheet.cell(row=2, column=col).value
        if review and word and word in review:
            sheet.cell(row=row, column=col, value=1)
        else:
            sheet.cell(row=row, column=col, value=0)

# Сохраняем книгу Excel
workbook.save('результат.xlsx')


