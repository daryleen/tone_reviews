from superSort import *
import openpyxl
import random
import pickle


def create_weights(all_unique_words):
    random.seed(2)
    random_weight = []
    for i in range(all_unique_words+1):
        kek = random.randint(1, 9)
        random_weight.append(kek)
    return random_weight


def evaluate_test_review(test_review, random_weights, border_net):
    net = 0
    for k in range(all_unique_words):
        net += random_weights[k] * test_review[k]  # Предполагается, что test_review - это список признаков для тестового отзыва

    if net > border_net:
        return "Положительный отзыв"
    else:
        return "Отрицательный отзыв"


# количество всех отзывов + кол-во уникальных слов для использования в циклах
all_reviews_len = len(positive_reviews) + len(negative_reviews)
all_unique_words = len(dataset_words)
# print(all_reviews_len,all_unique_words)
net = 0
random_weights = create_weights(all_unique_words)
# print(random_weights[0][0],"\n", random_weights[all_reviews_len-1])

# Загружаем Excel-файл
workbook = openpyxl.load_workbook("результат.xlsx")
sheet = workbook.active

# Считываем данные из ячеек A3 до A502
reviews = []
for row in sheet.iter_rows(min_row=3, max_row=all_reviews_len+2, min_col=1, max_col=1, values_only=True):
    reviews.append(row[0])

# Считываем значения из ячеек B3 до HH3 и записываем в массив arr
x_values = []
for row in sheet.iter_rows(min_row=3, max_row=all_reviews_len+2, min_col=2, max_col=all_unique_words+1, values_only=True):
    x_values.append(list(row))

test_review = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

EPOCH = 20000
border_net = 1000
count = 0
shuffled_indices = list(range(len(reviews)))
true_rev = []
false_rev = []
all_rev = [str(i) + ("T" if i < 250 else "F") for i in shuffled_indices]
# print(all_rev)
random.shuffle(all_rev)
max_correct_ans = -100
perfect_weights = []
for i in range(EPOCH):
    correct_ans = 0
    faults = 0
    # Перемешиваем индексы обучающих примеров перед каждой эпохой
    for j in range(len(all_rev)):
        net = 0
        for k in range(all_unique_words):
            #print(j,k)
            net += random_weights[k] * x_values[j][k]
        if ( "T" in all_rev[j] and net >= border_net) or ("F" in all_rev[j] and net < border_net):
            correct_ans += 1
        elif "T" in all_rev[j] and net < border_net:
            faults += 1
            for n in range(all_unique_words):
                if x_values[j][n] == 1:
                    random_weights[n] += 1
        elif "F" in all_rev[j] and net >= border_net:
            faults += 1
            for n in range(all_unique_words):
                if x_values[j][n] == 1:
                    random_weights[n] -= 1
    print(*random_weights)
    if correct_ans > max_correct_ans:
        max_correct_ans = correct_ans
        perfect_weights = random_weights
    correct_percent = correct_ans * 100 / all_reviews_len
    faults_percent = faults * 100 / all_reviews_len
    print(f"Epoch {i + 1}: Correct Answers: {correct_percent}%, Faults: {faults_percent}%")

# answer = evaluate_test_review(test_review, random_weights, border_net)
# print(answer)
print(perfect_weights)
print(max_correct_ans * 100 / all_reviews_len)

with open("perfect_weights.pkl", "wb") as file:
    pickle.dump(perfect_weights, file)

print("Переменная perfect_weights сохранена в файле perfect_weights.pkl")
